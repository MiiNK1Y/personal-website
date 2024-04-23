#!./.env/bin/python3

import os
import configparser
import openpyxl

 # THIS FIXES ANSI COLORS FOR WINDOW
from colorama import just_fix_windows_console
just_fix_windows_console()

class Excel_Workbook():
    def __init__(self, path: str) -> None:
        self.xlsx_file = path
        self.workbook = openpyxl.load_workbook(self.xlsx_file)
        self.sheet = self.workbook.active
        self.sheet_height = self.sheet.max_row
        self.sheet_width = self.sheet.max_column
        self.blank_cells = 0

    def create_sheet(self, sheet_name) -> None:
        self.workbook.create_sheet(sheet_name)

    def get_blank_cells(self) -> int:
        return self.blank_cells

    def get_all_column_headers(self) -> list:
        column_headers = []
        for header in self.sheet.iter_cols(1, self.sheet_width, values_only=True):
            column_headers.append(header[0])
        return column_headers

    def map_column_data_pair(self, column_0: int, column_1: int) -> list:
        column_data_pair = []
        rows = self.sheet.iter_rows(1, self.sheet_height, values_only=True)
        for row in rows:
            combined_value = str(row[column_0]) + "." + str(row[column_1])
            column_data_pair.append(combined_value)
        return column_data_pair

    def save_workbook(self, filename: str) -> None:
        self.workbook.save(filename)

    def close_workbook(self) -> None:
        self.workbook.close() 

    def delete_columns(self, columns_delete: list) -> None:
        column_index_compensation = 0
        for column in columns_delete:
            column -= column_index_compensation
            self.sheet.delete_cols(column)
            column_index_compensation += 1

    def match_and_insert(self, column_lane_index: int, column_insert_index: int, replacement_0: str, replacement_1: str, data: list) -> int:
        row_count = 1
        data_SNs = [i.split(".")[0] for i in data]
        rows = self.sheet.iter_rows(1, self.sheet_height, values_only=True)
        for row in rows:
            found_cell_value = row[column_lane_index]
            if found_cell_value == None:
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement_0)
                row_count += 1
                continue
            if found_cell_value in data_SNs:
                for value in data:
                    data_pair = value.split(".")
                    serial = data_pair[0]
                    po = data_pair[1]
                    if serial == found_cell_value:
                        self.sheet.cell(row=row_count, column=column_insert_index, value=po)
                        row_count += 1
            else:
                self.blank_cells += 1
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement_1)
                row_count += 1

def clr_stt(log_str: str) -> str:
    green = "\033[92m"
    cyan = "\033[36m"
    yellow = "\033[33m"
    red = "\033[41m"
    esc = "\033[0m"
    status = ""
    status_end_char = 0
    for index, i in enumerate(log_str):
        if i == log_str[0] and (log_str[0] == "["):
            status += i
            continue
        elif i == "]" and (index > 2 or index < 5):
            status += i
            status_end_char = index + 1
            break
        else:
            status += i
    rest_of_string = log_str[status_end_char:]
    rest_of_sequence = status + esc + rest_of_string
    if status == "[+]":
        return str(green + rest_of_sequence)
    elif status == "[x]":
        return str(cyan + rest_of_sequence)
    elif status == "[!]":
        return str(yellow + rest_of_sequence)
    elif status == "[!!]":
        return str(red + rest_of_sequence)

def date_is_valid(date: str) -> bool:
    days = int(date[0:2])
    months = int(date[3:5])
    years = int(date[6:])
    if (days <= 31) and (months <= 12) and ((years < 2100) and (years > 1970)) :
        return True
    else:
        return False

def get_date_from_str(file_name: str) -> str:
    valid_chars = ".0123456789" #instead of trying to convert every character to int
    date_format = "dd.mm.yyyy"
    for index, char in enumerate(file_name):
        if char in valid_chars:
            date_format_lenght = len(date_format) + 1
            for x in range(date_format_lenght):
                x_val = file_name[index + x]
                if (x == 2 or x == 5) and (x_val != "."):
                    break
                elif x == (date_format_lenght - 1):
                    date_found = file_name[index:(index + x)]
                    if date_is_valid(date_found):
                        return date_found
                    else:
                        break
                elif x_val not in valid_chars:
                    break
                else:
                    continue
        else:
            continue
    error_handler("Not able to get valid date from filename!")

# THIS ONE IS BIG BRAIN UNIQUE BRO.
def get_newest_date(d1: str, d2: str) -> str | None:
    if d1 == d2:
        error_handler("The dates in the filenames are the same, not able to decide which are the newest!")
    d1_formated = d1.split(".")[::-1]
    d2_formated = d2.split(".")[::-1]
    print(clr_stt("[+] Formated date: ") + str(d1_formated))
    print(clr_stt("[+] Formated date: ") + str(d2_formated))
    if d1_formated > d2_formated:
        print(clr_stt("[+] Newest date found: ") + d1)
        return d1
    else:
        print(clr_stt("[+] Newest date found: ") + d2)
        return d2

def get_excel_files() -> list:
    files = os.listdir(".")
    excel_files = []
    for file in files:
        if file[-5:] == ".xlsx":
            excel_files.append(file)
            print(clr_stt("[+] Found file: ") + file)
            if len(excel_files) > 2:
                error_handler("More than two excel files found in directory.")
    if len(excel_files) < 2:
        error_handler("Less than two excel files was found, two are needed for this operation.")
    return excel_files

def error_handler(msg) -> None:
    print("")
    print(clr_stt("[!!] Warning: ") + str(msg))
    print(clr_stt("[!!] Exit the application, fix the issue then retry."))
    print("")
    input(clr_stt("[x] Press [ENTER] to exit the application."))
    exit()

def main() -> None:
    print("\n----------logging started----------")

    # set configparser to read the config.
    config = configparser.ConfigParser()
    config.read("config.ini")
    default_config = config["DEFAULT"]

    # set all config.ini-values to their respective variables, making them usable.
    old_sheet_path = default_config["old_sheet_path"]
    new_sheet_path = default_config["new_sheet_path"]
    final_sheet_path = default_config["final_sheet_path"]
    final_sheet_name = default_config["final_sheet_name"]
    columns_to_delete = default_config["columns_to_delete"]
    serial_column_text = default_config["serial_column_text"]
    po_column_text = default_config["po_column_text"]
    final_sheet_column_insert_po = default_config["final_sheet_column_insert_po"]
    none_to_match_replacement = default_config["none_to_match_replacement"]
    no_match_replacement = default_config["no_match_replacement"]

    # converted to int, since configparser only keeps strings.
    final_sheet_column_insert_po = int(final_sheet_column_insert_po)

    # find the files to be manipulated.
    excel_files = get_excel_files()
    excel_file_1_date = get_date_from_str(excel_files[0])
    excel_file_2_date = get_date_from_str(excel_files[1])

    # assign the excel files to variables based on the dates detected in their names, NOT meta-data dates.
    newest_date = str(get_newest_date(excel_file_1_date, excel_file_2_date))
    if newest_date == excel_file_1_date:
        old_sheet_name = excel_files[1]
        new_sheet_name = excel_files[0]
    else:
        old_sheet_name = excel_files[0]
        new_sheet_name = excel_files[1]

    # assign the full file-path based on previous variables, to other variables that combine the path and the file-name.
    new_sheet_path = new_sheet_path + new_sheet_name
    print(clr_stt("[+] New sheet assigned to: ") + new_sheet_path)
    old_sheet_path = old_sheet_path + old_sheet_name
    print(clr_stt("[+] Old sheet assigned to: ") + old_sheet_path)
    final_sheet_path = final_sheet_path + final_sheet_name
    print("-----------logging ended-----------\n")

    # activate the workbooks.
    old_sheet = Excel_Workbook(old_sheet_path)
    new_sheet = Excel_Workbook(new_sheet_path)

    # need to convert "columns_to_delete" from string into list, for the compability.
    columns_to_delete = columns_to_delete.split(",")

    # use eval to convert every str-number to actual int.
    columns_to_delete = [eval(i) for i in columns_to_delete]
    new_sheet.delete_columns(columns_to_delete)
    new_sheet.save_workbook(final_sheet_path)
    new_sheet.close_workbook()

    final_sheet = Excel_Workbook(final_sheet_path)

    old_sheet_headers = old_sheet.get_all_column_headers()
    old_sheet_serial_index = old_sheet_headers.index(serial_column_text)
    old_sheet_po_index = old_sheet_headers.index(po_column_text)
    old_sheet_sn_po = old_sheet.map_column_data_pair(old_sheet_serial_index, old_sheet_po_index)
    final_sheet_headers = final_sheet.get_all_column_headers()
    final_sheet_serial_index = final_sheet_headers.index(serial_column_text)

    final_sheet.match_and_insert(
            column_lane_index=final_sheet_serial_index,
            column_insert_index=final_sheet_column_insert_po,
            replacement_0=none_to_match_replacement,
            replacement_1=no_match_replacement,
            data=old_sheet_sn_po
            )

    final_sheet_blank_POs = final_sheet.get_blank_cells()
    if final_sheet_blank_POs == 0:
        print(clr_stt("[+] Serial numbers without PO: ") + str(final_sheet_blank_POs))
        print(clr_stt("[+] No need to add new POs :)"))
    else:
        print(clr_stt("[!] Serial numbers without PO: ") + str(final_sheet_blank_POs))
        print(clr_stt("[!] Remember to add the new POs!!"))
        print(clr_stt(f"[!] '{no_match_replacement}' is written where the PO should be."))

    final_sheet.create_sheet("Available in stock")
    final_sheet.save_workbook(final_sheet_path)
    final_sheet.close_workbook()
    old_sheet.close_workbook()

    print("")
    print(clr_stt("[+] DONE!"))
    print("")
    input(clr_stt("[x] Press [ENTER] to exit the application."))

if __name__ == "__main__":
    main()